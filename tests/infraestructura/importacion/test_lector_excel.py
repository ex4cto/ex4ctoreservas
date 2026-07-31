from __future__ import annotations

import datetime
from pathlib import Path

import openpyxl
from openpyxl.worksheet.worksheet import Worksheet

from garay.dominio.comun.dinero import Dinero
from garay.infraestructura.importacion.lector_excel import LectorVentasExcelOpenpyxl

_BASE = [
    "Fecha de servicio",
    "Cliente",
    "Cerrador",
    "Vendedor",
    "Descripción",
    "Costo Neto Total",
    "Valor total",
    "Margen de Ganancia total",
]


def _escribir(ws: Worksheet, titulo: str, headers: list[str], filas: list[list[object]]) -> None:
    ws.append([titulo])  # fila 1 (titulo)
    ws.append([])  # fila 2 (en blanco)
    ws.append(headers)  # fila 3 (encabezados)
    for f in filas:
        ws.append(f)


def _crear_xlsx(ruta: Path) -> None:
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    _escribir(
        wb.create_sheet("Hotel"),
        "RESERVAS HOTEL",
        [*_BASE, "Agencia 60%"],
        [[datetime.date(2026, 7, 5), "Gonzalo Lopez", "Eduardo", "Eduardo",
          "City tour", 660000, 880000, 220000, 132000]],
    )
    _escribir(
        wb.create_sheet("Externas"),
        "RESERVAS EXTERNAS",
        [*_BASE, "Agencia 40%"],
        [[datetime.date(2026, 7, 3), "Stefan Ponce", "Mairele", "Mairele",
          "Playa tranquila", 300000, 400000, 120000, 48000]],
    )
    _escribir(
        wb.create_sheet(" Punto crespo 50"),
        "RESERVAS PUNTO CRESPO",
        [*_BASE, "Agencia 30%"],
        [
            # compartida con Externas -> debe deduplicarse
            [datetime.date(2026, 7, 3), "Stefan Ponce", "Mairele", "Mairele",
             "Playa tranquila", 300000, 400000, 120000, 36000],
            # exclusiva de Crespo
            [datetime.date(2026, 7, 18), "Sandy Martinez", "Hebert", "Hebert",
             "5 islas", 380000, 560000, 180000, 54000],
        ],
    )
    _escribir(
        wb.create_sheet("Punto Crespo 60"),
        "RESERVAS PUNTO CRESPO",
        [*_BASE, "Agencia 20%"],
        [  # misma exclusiva de Crespo -> se deduplica dentro de la familia PC
            [datetime.date(2026, 7, 18), "Sandy Martinez", "Hebert", "Hebert",
             "5 islas", 380000, 560000, 180000, 36000]],
    )
    wb.save(ruta)


def test_lee_y_deduplica(tmp_path: Path) -> None:
    ruta = tmp_path / "conta.xlsx"
    _crear_xlsx(ruta)
    filas = LectorVentasExcelOpenpyxl().leer(str(ruta))

    # Hotel(1) + Externas(1) + Crespo exclusivo(1) = 3; el compartido se deduplica
    assert len(filas) == 3
    assert {f.canal for f in filas} == {"Hotel", "Externas", "Crespo"}


def test_campos_y_agencia_por_hoja(tmp_path: Path) -> None:
    ruta = tmp_path / "conta.xlsx"
    _crear_xlsx(ruta)
    filas = {f.canal: f for f in LectorVentasExcelOpenpyxl().leer(str(ruta))}

    hotel = filas["Hotel"]
    assert hotel.cliente_nombre == "Gonzalo Lopez"
    assert hotel.fecha == datetime.date(2026, 7, 5)
    assert hotel.valor == Dinero("880000")
    assert hotel.neto == Dinero("660000")
    assert hotel.margen == Dinero("220000")
    assert hotel.agencia == Dinero("132000")  # columna Agencia 60%

    crespo = filas["Crespo"]
    assert crespo.cliente_nombre == "Sandy Martinez"
    assert crespo.agencia == Dinero("54000")  # de PC50 (Agencia 30%), gana el primero


def test_lee_montos_de_comision(tmp_path: Path) -> None:
    ruta = tmp_path / "conta.xlsx"
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    _escribir(
        wb.create_sheet("Hotel"),
        "RESERVAS HOTEL",
        [*_BASE, "Agencia 60%", "Vendedor 20%", "Cerrador 20%"],
        [[datetime.date(2026, 7, 5), "Gonzalo Lopez", "Eduardo", "Eduardo",
          "City tour", 660000, 880000, 220000, 132000, 44000, 44000]],
    )
    wb.save(ruta)
    fila = LectorVentasExcelOpenpyxl().leer(str(ruta))[0]
    assert fila.comision_vendedor == Dinero("44000")
    assert fila.comision_cerrador == Dinero("44000")
    # la columna de NOMBRE "Vendedor" no se confunde con el monto "Vendedor 20%"
    assert fila.vendedor_nombre == "Eduardo"


def test_lee_fecha_guardada_como_texto(tmp_path: Path) -> None:
    # En la planilla real algunas fechas quedan como string; deben leerse igual.
    ruta = tmp_path / "conta.xlsx"
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    _escribir(
        wb.create_sheet("Hotel"),
        "RESERVAS HOTEL",
        [*_BASE, "Agencia 60%"],
        [["2026-07-09", "Ana Ruiz", "X", "Y", "City tour", 100000, 200000, 100000, 60000]],
    )
    wb.save(ruta)
    filas = LectorVentasExcelOpenpyxl().leer(str(ruta))
    assert len(filas) == 1
    assert filas[0].fecha == datetime.date(2026, 7, 9)


def test_salta_filas_sin_valor(tmp_path: Path) -> None:
    ruta = tmp_path / "conta.xlsx"
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    _escribir(
        wb.create_sheet("Hotel"),
        "RESERVAS HOTEL",
        [*_BASE, "Agencia 60%"],
        [
            [datetime.date(2026, 7, 4), "SinVal", "X", "Y", "Tour", 100000, None, 0, 0],
            [datetime.date(2026, 7, 5), "ConVal", "X", "Y", "Tour", 100000, 200000, 100000, 60000],
        ],
    )
    wb.save(ruta)
    filas = LectorVentasExcelOpenpyxl().leer(str(ruta))
    assert len(filas) == 1
    assert filas[0].cliente_nombre == "ConVal"
