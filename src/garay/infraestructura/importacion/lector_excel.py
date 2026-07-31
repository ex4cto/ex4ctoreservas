"""Adaptador que lee la planilla de contabilidad de Garay con openpyxl.

La estructura de cada hoja de ventas: fila 1 titulo, fila 2 en blanco, fila 3
encabezados, datos desde la fila 4. El canal de venta es la identidad de la hoja
(Hotel / Externas / Crespo). Las hojas de Crespo (PC50 y PC60) son dos escenarios
de comision de las MISMAS ventas: se deduplican entre si, y las que ya estan en
Externas se descartan (Externas es el maestro; solo las exclusivas quedan como Crespo).
"""

from __future__ import annotations

import datetime
import unicodedata
from decimal import Decimal, InvalidOperation

from openpyxl import load_workbook
from openpyxl.workbook.workbook import Workbook

from garay.dominio.comun.dinero import Dinero
from garay.dominio.puertos.servicios_externos import FilaVentaImportada, LectorVentasExcel

_HEADER_ROW = 2  # indice 0-based: fila 1 titulo, fila 2 blanco, fila 3 encabezados

# (nombre de hoja, canal de venta, encabezado de la columna de agencia)
_HOJAS: list[tuple[str, str, str]] = [
    ("Hotel", "Hotel", "Agencia 60%"),
    ("Externas", "Externas", "Agencia 40%"),
    (" Punto crespo 50", "Crespo", "Agencia 30%"),
    ("Punto Crespo 60", "Crespo", "Agencia 20%"),
]


def _norm(texto: object) -> str:
    s = unicodedata.normalize("NFKD", str(texto)).encode("ascii", "ignore").decode()
    return " ".join(s.strip().lower().split())


def _a_dinero(valor: object) -> Dinero | None:
    if valor is None or (isinstance(valor, str) and not valor.strip()):
        return None
    try:
        return Dinero(Decimal(str(valor)))
    except (InvalidOperation, ValueError):
        return None


def _a_fecha(valor: object) -> datetime.date | None:
    if isinstance(valor, datetime.datetime):
        return valor.date()
    if isinstance(valor, datetime.date):
        return valor
    if isinstance(valor, str):
        texto = valor.strip()
        try:
            return datetime.date.fromisoformat(texto[:10])
        except ValueError:
            pass
        for fmt in ("%d/%m/%Y", "%d/%m/%y", "%m/%d/%Y"):
            try:
                return datetime.datetime.strptime(texto, fmt).date()
            except ValueError:
                continue
    return None


def _texto(valor: object) -> str:
    return str(valor).strip() if valor is not None else ""


def _clave(f: FilaVentaImportada) -> tuple[str, str, int, str]:
    return (
        f.cliente_nombre.lower(),
        f.fecha.isoformat(),
        int(f.valor.monto),
        f.descripcion.lower()[:15],
    )


class LectorVentasExcelOpenpyxl(LectorVentasExcel):
    def leer(self, ruta: str) -> list[FilaVentaImportada]:
        wb = load_workbook(ruta, data_only=True, read_only=True)
        try:
            por_hoja = {
                nombre: self._leer_hoja(wb, nombre, canal, agencia)
                for nombre, canal, agencia in _HOJAS
            }
            hotel = por_hoja["Hotel"]
            externas = por_hoja["Externas"]
            crespo_todos = _dedup(por_hoja[" Punto crespo 50"] + por_hoja["Punto Crespo 60"])
            ext_keys = {_clave(f) for f in externas}
            crespo = [f for f in crespo_todos if _clave(f) not in ext_keys]
            return hotel + externas + crespo
        finally:
            wb.close()

    def _leer_hoja(
        self, wb: Workbook, nombre: str, canal: str, agencia_header: str
    ) -> list[FilaVentaImportada]:
        if nombre not in wb.sheetnames:
            return []
        filas_crudas = list(wb[nombre].iter_rows(values_only=True))
        if len(filas_crudas) <= _HEADER_ROW:
            return []
        encabezados = filas_crudas[_HEADER_ROW]
        idx = {_norm(h): i for i, h in enumerate(encabezados) if h is not None}
        # Columnas de MONTO de comision: "Vendedor 20%", "Cerrador 50%"... (el % varia
        # por hoja). Se distinguen de las columnas de NOMBRE ("Vendedor", "Cerrador")
        # porque llevan algo mas despues del prefijo.
        i_com_vend = next((i for k, i in idx.items() if k.startswith("vendedor ")), None)
        i_com_cerr = next((i for k, i in idx.items() if k.startswith("cerrador ")), None)

        def col(fila: tuple[object, ...], nombre_col: str) -> object:
            i = idx.get(_norm(nombre_col))
            return fila[i] if i is not None and i < len(fila) else None

        def por_indice(fila: tuple[object, ...], i: int | None) -> object:
            return fila[i] if i is not None and i < len(fila) else None

        resultado: list[FilaVentaImportada] = []
        for fila in filas_crudas[_HEADER_ROW + 1 :]:
            fecha = _a_fecha(col(fila, "Fecha de servicio"))
            valor = _a_dinero(col(fila, "Valor total"))
            cliente = _texto(col(fila, "Cliente"))
            if fecha is None or valor is None or not cliente:
                continue
            neto = _a_dinero(col(fila, "Costo Neto Total")) or Dinero("0")
            margen = _a_dinero(col(fila, "Margen de Ganancia total")) or (valor - neto)
            agencia = _a_dinero(col(fila, agencia_header)) or Dinero("0")
            resultado.append(
                FilaVentaImportada(
                    canal=canal,
                    cliente_nombre=cliente,
                    fecha=fecha,
                    descripcion=_texto(col(fila, "Descripción")),
                    valor=valor,
                    neto=neto,
                    margen=margen,
                    agencia=agencia,
                    comision_vendedor=_a_dinero(por_indice(fila, i_com_vend)) or Dinero("0"),
                    comision_cerrador=_a_dinero(por_indice(fila, i_com_cerr)) or Dinero("0"),
                    vendedor_nombre=_texto(col(fila, "Vendedor")) or None,
                    cerrador_nombre=_texto(col(fila, "Cerrador")) or None,
                )
            )
        return resultado


def _dedup(filas: list[FilaVentaImportada]) -> list[FilaVentaImportada]:
    vistas: set[tuple[str, str, int, str]] = set()
    salida: list[FilaVentaImportada] = []
    for f in filas:
        k = _clave(f)
        if k not in vistas:
            vistas.add(k)
            salida.append(f)
    return salida
